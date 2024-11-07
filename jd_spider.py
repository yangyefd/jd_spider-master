import tkinter as tk
from tkinter import filedialog
from openpyxl import load_workbook
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import time
import tkinter.ttk as ttk
from threading import Thread

def format_chinese_visual(text, width=30):
    # 定义一个函数，用于计算字符串的视觉宽度
    def visual_width(s):
        return sum(2 if '\u4e00' <= c <= '\u9fff' else 1 for c in s)

    # 如果字符串的视觉宽度大于指定的宽度，则截断字符串
    if visual_width(text) > width:
        final_text = ""
        current_width = 0
        for char in text:
            if current_width + (2 if '\u4e00' <= char <= '\u9fff' else 1) > width:
                break
            final_text += char
            current_width += 2 if '\u4e00' <= char <= '\u9fff' else 1
    else:
        final_text = text

    # 使用空格填充剩余的宽度
    return final_text + ' ' * (width - visual_width(final_text))

def jd_search_selenium(keyword, target_keyword, idx, text_widget, driver):
    # 打开搜索页面
    driver.get(f"https://search.jd.com/Search?keyword={keyword}&enc=utf-8")

    try:
        # 等待商品列表元素加载完成
        wait = WebDriverWait(driver, 50)
        product_list = wait.until(EC.presence_of_all_elements_located((By.XPATH, '//div[@class="gl-i-wrap"]')))
        
        result = 0
        for item in product_list:
            # 尝试使用 XPATH 选择器找到商品链接
            try:
                link = item.find_element(By.XPATH, './/div[@class="p-name p-name-type-2"]/a')
                product_url = link.get_attribute('href')
                
                print('product_url:', product_url)
                if target_keyword in product_url:
                    result = 1
                    break
                
                time.sleep(1)  # 适当休眠，避免被识别为爬虫
            except Exception as e:
                print(f"Error finding link: {e}")

        # 在Text组件中显示结果
        if result == 1:
            text_widget.insert(tk.END, f"{str(idx).ljust(4)} | {format_chinese_visual(keyword)} | {target_keyword.ljust(20):<20} | 找到商品\n")
        else:
            text_widget.insert(tk.END, f"{str(idx).ljust(4)} | {format_chinese_visual(keyword)} | {target_keyword.ljust(20):<20} | 未找到商品\n", "red_text")

        
        return result
    except Exception as e:
        print(f"An error occurred: {e}")
        return []

def read_excel(file_path, column_title="商品关键词"):
    target_SPUID = "所属商品SPUID"
    # 读取Excel文件
    wb = load_workbook(file_path)
    sheet = wb.active

    # 找到关键词所在的列
    header_row = sheet[1]  # 假设第一行是标题行
    keyword_column_index = None
    SPUID_column_index = None
    for cell in header_row:
        if cell.value == column_title:
            keyword_column_index = cell.column
            break
    for cell in header_row:
        if cell.value == target_SPUID:
            SPUID_column_index = cell.column
            break

    if keyword_column_index is None or SPUID_column_index is None:
        raise ValueError(f"Column with title '{column_title}' not found.")

    idxs = []
    keywords = []
    SPUIDs = []
    idx = 2
    for row in sheet.iter_rows(min_row=2, values_only=True):
        # 根据找到的列索引读取关键词
        keyword = row[header_row.index(next(cell for cell in header_row if cell.column == keyword_column_index))]
        SPUID = row[header_row.index(next(cell for cell in header_row if cell.column == SPUID_column_index))]
        keywords.append(keyword)
        SPUIDs.append(SPUID)
        idxs.append(idx)
        idx += 1

    return keywords, SPUIDs, idxs

def search_from_excel(text_widget, driver, progress ):
    file_path = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx")])
    if not file_path:
        return
    
    keywords, SPUIDs, idxs = read_excel(file_path)
    progress['maximum'] = len(keywords)  # 设置进度条的最大值
    for keyword, target_keyword, idx in zip(keywords, SPUIDs, idxs):
        jd_search_selenium(keyword, target_keyword, idx, text_widget, driver)
        progress['value'] += 1
        root.update_idletasks()

def start_search_thread(text_results, driver, progress):
    # 创建并启动线程
    search_thread = Thread(target=search_from_excel, args=(text_results, driver, progress))
    search_thread.start()

# 创建Tkinter窗口
root = tk.Tk()
root.title("京东商品搜索")

# 使用Grid布局管理器，并设置行列权重
root.grid_rowconfigure(1, weight=1)
root.grid_columnconfigure(0, weight=1)

# 创建Text组件用于显示信息
text_results = tk.Text(root, height=20, width=100)
text_results.grid(row=1, column=0, sticky="nsew")  # 使用sticky="nsew"使组件在所有方向上都可以扩展
# text_results.tag_configure("TITLE", font=("Arial", 10, "bold"), foreground="blue")
text_results.insert(tk.END, f"{'idx'.ljust(4)} | {'keywords'.ljust(30)} | {'SPU'.ljust(20)} | {'results'.ljust(15)}\n")
text_results.tag_configure("red_text", foreground="red")

# 创建浏览器实例
driver = webdriver.Edge()

# 创建进度条
progress = ttk.Progressbar(root, orient="horizontal", length=200, mode="determinate")
progress.grid(row=2, column=0, sticky="ew")  # 使用sticky="ew"使进度条在水平方向上可以扩展

# 创建按钮
button_search = tk.Button(root, text="选择Excel文件并搜索", command=lambda: start_search_thread(text_results, driver, progress))
button_search.grid(row=0, column=0, sticky="ew")  # 使用sticky="ew"使按钮在水平方向上可以扩展

# 当窗口关闭时，关闭浏览器
root.protocol("WM_DELETE_WINDOW", lambda: [root.destroy(), driver.quit()])

# 运行Tkinter事件循环
root.mainloop()